from flask import Flask, render_template, request, redirect, flash, url_for, session, send_file, jsonify, Blueprint
import sqlite3
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
import os
from openpyxl import Workbook
import shutil
import io
from datetime import datetime
import pandas as pd
import os
from models import db
from models.santri import Santri

app = Flask(__name__)
import os
print(">> DATABASE YANG DIPAKAI:", os.path.abspath("erp_pondok.db"))

app.secret_key = 'rahasia'
UPLOAD_FOLDER = 'static/uploads/'
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'xls', 'xlsx', 'csv'}
bp = Blueprint('santri', __name__)  # ganti sesuai struktur projek Anda

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///erp_pondok.db'
db.init_app(app)

def format_uang(value):
    return f"{value:,.0f}".replace(",", ".")

# Daftarkan sebagai filter Jinja
app.jinja_env.filters['format_uang'] = format_uang

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def get_db_connection():
    conn = sqlite3.connect('erp_pondok.db')
    conn.row_factory = sqlite3.Row
    return conn

def auto_backup_harian():
    try:
        backup_folder = 'backup'
        log_file = os.path.join(backup_folder, 'log_backup.txt')

        if not os.path.exists(backup_folder):
            os.makedirs(backup_folder)

        today = datetime.now().strftime('%Y%m%d')
        backup_filename = f'backup_{today}.db'
        backup_path = os.path.join(backup_folder, backup_filename)

        # Cek apakah backup hari ini sudah ada
        if not os.path.exists(backup_path):
            shutil.copy2('erp_pondok.db', backup_path)
            log_message = f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] Backup berhasil: {backup_filename}\n"
        else:
            log_message = f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] Backup sudah ada: {backup_filename}\n"

    except Exception as e:
        log_message = f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] Backup gagal: {str(e)}\n"

    # Simpan log ke file
    with open(log_file, 'a') as f:
        f.write(log_message)

# Fungsi koneksi ke database
def get_db_connection():
    conn = sqlite3.connect('erp_pondok.db')
    conn.row_factory = sqlite3.Row
    return conn

# Fungsi ambil settings dari database
def load_settings():
    conn = get_db_connection()
    setting = conn.execute('SELECT * FROM settings WHERE id = 1').fetchone()
    conn.close()
    return setting

# Fungsi simpan settings ke database
def save_settings(data):
    conn = get_db_connection()
    if conn.execute('SELECT COUNT(*) FROM settings WHERE id = 1').fetchone()[0] == 0:
        conn.execute('INSERT INTO settings (id) VALUES (1)')
    conn.execute('''
        UPDATE settings SET
            nama_pondok = ?,
            alamat = ?,
            kepala_pondok = ?,
            moto = ?,
            logo = ?,
            tema = ?,
            ukuran_upload = ?,
            format_santri = ?,
            pagination = ?,
            tahun_ajaran = ?,
            semester_mulai = ?,
            semester_selesai = ?,
            aktifkan_notif = ?,
            api_fonnte_1 = ?,
            api_fonnte_2 = ?,
            jadwal_kirim = ?,
            template_pesan = ?,
            jadwal_backup = ?,
            durasi_backup = ?,
            timeout = ?
        WHERE id = 1
    ''', (
        data.get('nama_pondok'),
        data.get('alamat'),
        data.get('kepala_pondok'),
        data.get('moto'),
        data.get('logo'),
        data.get('tema'),
        data.get('ukuran_upload'),
        data.get('format_santri'),
        data.get('pagination'),
        data.get('tahun_ajaran'),
        data.get('semester_mulai'),
        data.get('semester_selesai'),
        data.get('aktifkan_notif'),
        data.get('api_fonnte_1'),
        data.get('api_fonnte_2'),
        data.get('jadwal_kirim'),
        data.get('template_pesan'),
        data.get('jadwal_backup'),
        data.get('durasi_backup'),
        data.get('timeout')
    ))
    conn.commit()
    conn.close()


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@bp.route('/import-santri', methods=['GET', 'POST'])
def import_santri():
    if request.method == 'POST':
        file = request.files.get('file')
        if not file or not allowed_file(file.filename):
            flash("Format file tidak didukung. Gunakan .xls, .xlsx, atau .csv.", 'danger')
            return redirect(request.url)

        filename = secure_filename(file.filename)
        filepath = os.path.join('uploads', filename)
        file.save(filepath)

        # Baca file dengan pandas
        try:
            if filename.endswith('.csv'):
                df = pd.read_csv(filepath)
            else:
                df = pd.read_excel(filepath)

            # Pastikan kolom yang dibutuhkan ada
            required_columns = {'NIS', 'Nama', 'Kelas', 'Jenis Kelamin', 'Tanggal Lahir', 'Alamat'}
            if not required_columns.issubset(df.columns):
                flash("Kolom pada file tidak sesuai template.", 'danger')
                return redirect(request.url)

            # Simpan ke database (contoh)
            for _, row in df.iterrows():
                santri = Santri(
                    nis=row['NIS'],
                    nama=row['Nama'],
                    kelas=row['Kelas'],
                    jenis_kelamin=row['Jenis Kelamin'],
                    tanggal_lahir=row['Tanggal Lahir'],
                    alamat=row['Alamat']
                )
                db.session.add(santri)

            db.session.commit()
            flash("Data santri berhasil diimpor.", 'success')
        except Exception as e:
            flash(f"Gagal mengimpor: {str(e)}", 'danger')
        finally:
            os.remove(filepath)

        return redirect(url_for('dashboard'))

    return render_template('import_santri.html')

# Home Page
@app.route('/')
def home():
    return render_template('home.html')

# Login Page
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password_input = request.form['password']

        conn = get_db_connection()
        user = conn.execute('SELECT * FROM user WHERE username = ? AND is_active = 1', (username,)).fetchone()
        conn.close()

        if user and check_password_hash(user['password'], password_input):
            session['id_user'] = user['id_user']
            session['username'] = user['username']
            session['nama_lengkap'] = user['nama_lengkap']
            session['role'] = user['role']
            session['logged_in'] = True  

            conn = get_db_connection()
            conn.execute('UPDATE user SET last_login = datetime("now") WHERE id_user = ?', (user['id_user'],))
            conn.commit()
            conn.close()

            flash('Login berhasil!', 'success')
            return redirect(url_for('dashboard'))

        if user['role'] == 'admin':
           auto_backup_harian()
        else:
            flash('Username atau password salah.', 'danger')

    return render_template("login.html", body_class="login-page")

# Dashboard Page
@app.route('/dashboard')
def dashboard():
    if 'username' not in session:
        return redirect(url_for('login'))

    role = session.get('role')

    if role == 'admin':
        conn = get_db_connection()
        total_bulan_ini = conn.execute('''
            SELECT IFNULL(SUM(nominal_dibayar),0) FROM pembayaran 
            WHERE bulan = strftime('%m', 'now') AND tahun = strftime('%Y', 'now')
        ''').fetchone()[0]

        tunggakan_bulan_ini = conn.execute('''
            SELECT IFNULL(SUM(nominal_tagihan - nominal_dibayar),0) FROM pembayaran 
            WHERE status_bayar = 'Belum Lunas' AND bulan = strftime('%m', 'now') AND tahun = strftime('%Y', 'now')
        ''').fetchone()[0]

        total_semua = conn.execute('''
            SELECT IFNULL(SUM(nominal_dibayar),0) FROM pembayaran
        ''').fetchone()[0]

        tunggakan_semua = conn.execute('''
            SELECT IFNULL(SUM(nominal_tagihan - nominal_dibayar),0) FROM pembayaran
            WHERE status_bayar = 'Belum Lunas'
        ''').fetchone()[0]

        # Data untuk grafik pemasukan per bulan tahun berjalan
        data_grafik = conn.execute('''
            SELECT bulan, SUM(nominal_dibayar) as total_bayar
            FROM pembayaran
            WHERE tahun = strftime('%Y', 'now')
            GROUP BY bulan
            ORDER BY CASE
                WHEN bulan = 'Januari' THEN 1
                WHEN bulan = 'Februari' THEN 2
                WHEN bulan = 'Maret' THEN 3
                WHEN bulan = 'April' THEN 4
                WHEN bulan = 'Mei' THEN 5
                WHEN bulan = 'Juni' THEN 6
                WHEN bulan = 'Juli' THEN 7
                WHEN bulan = 'Agustus' THEN 8
                WHEN bulan = 'September' THEN 9
                WHEN bulan = 'Oktober' THEN 10
                WHEN bulan = 'November' THEN 11
                WHEN bulan = 'Desember' THEN 12
            END
        ''').fetchall()
        conn.close()

        bulan_grafik = [row['bulan'] for row in data_grafik]
        total_grafik = [row['total_bayar'] for row in data_grafik]

        return render_template('dashboard_admin.html',
                       total_bulan_ini=total_bulan_ini,
                       tunggakan_bulan_ini=tunggakan_bulan_ini,
                       total_semua=total_semua,
                       tunggakan_semua=tunggakan_semua,
                       bulan_grafik=bulan_grafik,
                       total_grafik=total_grafik,
                       body_class='logged-in')

    elif role == 'bendahara':
        conn = get_db_connection()
        total_bulan_ini = conn.execute('''
            SELECT IFNULL(SUM(nominal_dibayar),0) FROM pembayaran 
            WHERE bulan = strftime('%m', 'now') AND tahun = strftime('%Y', 'now')
        ''').fetchone()[0]

        tunggakan_bulan_ini = conn.execute('''
            SELECT IFNULL(SUM(nominal_tagihan - nominal_dibayar),0) FROM pembayaran 
            WHERE status_bayar = 'Belum Lunas' AND bulan = strftime('%m', 'now') AND tahun = strftime('%Y', 'now')
        ''').fetchone()[0]

        total_semua = conn.execute('''
            SELECT IFNULL(SUM(nominal_dibayar),0) FROM pembayaran
        ''').fetchone()[0]

        data_grafik = conn.execute('''
            SELECT bulan, SUM(nominal_dibayar) as total_bayar
            FROM pembayaran
            WHERE tahun = strftime('%Y', 'now')
            GROUP BY bulan
            ORDER BY CASE
                WHEN bulan = 'Januari' THEN 1
                WHEN bulan = 'Februari' THEN 2
                WHEN bulan = 'Maret' THEN 3
                WHEN bulan = 'April' THEN 4
                WHEN bulan = 'Mei' THEN 5
                WHEN bulan = 'Juni' THEN 6
                WHEN bulan = 'Juli' THEN 7
                WHEN bulan = 'Agustus' THEN 8
                WHEN bulan = 'September' THEN 9
                WHEN bulan = 'Oktober' THEN 10
                WHEN bulan = 'November' THEN 11
                WHEN bulan = 'Desember' THEN 12
            END
        ''').fetchall()
        conn.close()

        bulan_grafik = [row['bulan'] for row in data_grafik]
        total_grafik = [row['total_bayar'] for row in data_grafik]

        return render_template('dashboard_bendahara.html',
                           total_bulan_ini=total_bulan_ini,
                           tunggakan_bulan_ini=tunggakan_bulan_ini,
                           total_semua=total_semua,
                           bulan_grafik=bulan_grafik,
                           total_grafik=total_grafik)

    elif role == 'keuangan':
        return render_template('dashboard_keuangan.html', body_class='logged-in')

    else:
        flash('Role tidak dikenali.', 'danger')
        return redirect(url_for('login'))

# Logout
@app.route('/logout')
def logout():
    session.clear()
    flash('Berhasil logout.', 'info')
    return redirect(url_for('login'))

# Tambah User
@app.route('/tambah_user', methods=['GET', 'POST'])
def tambah_user():
    if request.method == 'POST':
        username = request.form['username']
        password = generate_password_hash(request.form['password'])
        nama_lengkap = request.form['nama_lengkap']
        role = request.form['role']
        email = request.form['email']
        no_hp = request.form['no_hp']

        conn = get_db_connection()
        conn.execute('''
            INSERT INTO user (username, password, nama_lengkap, role, email, no_hp, is_active) 
            VALUES (?, ?, ?, ?, ?, ?, 1)
        ''', (username, password, nama_lengkap, role, email, no_hp))
        conn.commit()
        conn.close()

        flash('User baru berhasil ditambahkan!', 'success')
        return redirect('/tambah_user')

    return render_template('tambah_user.html')

# List User
@app.route('/list_user')
def list_user():
    conn = get_db_connection()
    users = conn.execute('SELECT * FROM user').fetchall()
    conn.close()
    return render_template('list_user.html', users=users)

# Hapus User
@app.route('/hapus_user/<int:id>')
def hapus_user(id):
    conn = get_db_connection()
    conn.execute('DELETE FROM user WHERE id_user = ?', (id,))
    conn.commit()
    conn.close()
    flash('User berhasil dihapus!', 'success')
    return redirect('/list_user')

@app.route('/tambah_santri', methods=['GET', 'POST'])
def tambah_santri():
    if 'username' not in session:
        return redirect(url_for('login'))

    if request.method == 'POST':
        nis = request.form['nis']
        nama_lengkap = request.form['nama_lengkap']
        tempat_lahir = request.form['tempat_lahir']
        tanggal_lahir = request.form['tanggal_lahir']
        jenis_kelamin = request.form['jenis_kelamin']
        alamat = request.form['alamat']
        no_hp = request.form['no_hp']
        wali_santri = request.form['wali_santri']
        kelas = request.form['kelas']
        kamar = request.form['kamar']
        status = request.form['status']
        tanggal_masuk = request.form['tanggal_masuk']
        foto_filename = None

        if 'foto' in request.files:
            foto = request.files['foto']
            if foto and allowed_file(foto.filename):
                filename = secure_filename(foto.filename)
                foto.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                foto_filename = filename

        conn = get_db_connection()
        conn.execute('''
            INSERT INTO santri (nis, nama_lengkap, tempat_lahir, tanggal_lahir, jenis_kelamin, alamat, no_hp, wali_santri, kelas, kamar, status, tanggal_masuk, foto)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (nis, nama_lengkap, tempat_lahir, tanggal_lahir, jenis_kelamin, alamat, no_hp, wali_santri, kelas, kamar, status, tanggal_masuk, foto_filename))
        conn.commit()
        conn.close()

        flash('Data santri berhasil disimpan!', 'success')
        return redirect(url_for('tambah_santri'))

    return render_template('tambah_santri.html')

# List Santri
@app.route('/list_santri')
def list_santri():
    if 'username' not in session:
        return redirect(url_for('login'))

    q = request.args.get('q')
    conn = get_db_connection()

    if q:
        query = f"""
        SELECT * FROM santri
        WHERE nama_lengkap LIKE ? OR nis LIKE ?
        """
        santri_list = conn.execute(query, (f'%{q}%', f'%{q}%')).fetchall()
    else:
        santri_list = conn.execute('SELECT * FROM santri').fetchall()

    conn.close()
    return render_template('list_santri.html', santri_list=santri_list)

# Hapus Santri
@app.route('/hapus_santri/<int:id>')
def hapus_santri(id):
    if 'username' not in session:
        return redirect(url_for('login'))

    conn = get_db_connection()
    conn.execute('DELETE FROM santri WHERE id_santri = ?', (id,))
    conn.commit()
    conn.close()

    flash('Data santri berhasil dihapus!', 'success')
    return redirect(url_for('list_santri'))

@app.route('/edit_santri/<int:id>', methods=['GET', 'POST'])
def edit_santri(id):
    if 'username' not in session:
        return redirect(url_for('login'))

    conn = get_db_connection()
    santri = conn.execute('SELECT * FROM santri WHERE id_santri = ?', (id,)).fetchone()

    if request.method == 'POST':
        nis = request.form['nis']
        nama_lengkap = request.form['nama_lengkap']
        tempat_lahir = request.form['tempat_lahir']
        tanggal_lahir = request.form['tanggal_lahir']
        jenis_kelamin = request.form['jenis_kelamin']
        alamat = request.form['alamat']
        no_hp = request.form['no_hp']
        wali_santri = request.form['wali_santri']
        kelas = request.form['kelas']
        kamar = request.form['kamar']
        status = request.form['status']
        tanggal_masuk = request.form['tanggal_masuk']
        foto_filename = santri['foto']

        if 'foto' in request.files:
            foto = request.files['foto']
            if foto and allowed_file(foto.filename):
                filename = secure_filename(foto.filename)
                foto.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                foto_filename = filename

        conn.execute('''
            UPDATE santri SET
                nis = ?, nama_lengkap = ?, tempat_lahir = ?, tanggal_lahir = ?,
                jenis_kelamin = ?, alamat = ?, no_hp = ?, wali_santri = ?,
                kelas = ?, kamar = ?, status = ?, tanggal_masuk = ?, foto = ?
            WHERE id_santri = ?
        ''', (nis, nama_lengkap, tempat_lahir, tanggal_lahir, jenis_kelamin, alamat, no_hp, wali_santri, kelas, kamar, status, tanggal_masuk, foto_filename, id))

        conn.commit()
        conn.close()

        flash('Data santri berhasil diperbarui!', 'success')
        return redirect(url_for('list_santri'))

    conn.close()
    return render_template('edit_santri.html', santri=santri)

@app.route('/detail_santri/<int:id>')
def detail_santri(id):
    if 'username' not in session:
        return redirect(url_for('login'))

    conn = get_db_connection()
    santri = conn.execute('SELECT * FROM santri WHERE id_santri = ?', (id,)).fetchone()
    conn.close()

    if not santri:
        flash('Santri tidak ditemukan.', 'danger')
        return redirect(url_for('list_santri'))

    return render_template('detail_santri.html', santri=santri)

# Route Export Santri
@app.route('/export_santri')
def export_santri():
    if 'username' not in session:
        return redirect(url_for('login'))

    conn = get_db_connection()
    santri_list = conn.execute('SELECT * FROM santri').fetchall()
    conn.close()

    # Buat workbook Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Data Santri"

    # Header kolom
    headers = ['NIS', 'Nama Lengkap', 'Tempat Lahir', 'Tanggal Lahir', 'Jenis Kelamin', 'Alamat', 'No HP', 'Wali Santri', 'Kelas', 'Kamar', 'Status', 'Tanggal Masuk']
    ws.append(headers)

    # Isi data
    for santri in santri_list:
        ws.append([
            santri['nis'],
            santri['nama_lengkap'],
            santri['tempat_lahir'],
            santri['tanggal_lahir'],
            santri['jenis_kelamin'],
            santri['alamat'],
            santri['no_hp'],
            santri['wali_santri'],
            santri['kelas'],
            santri['kamar'],
            santri['status'],
            santri['tanggal_masuk']
        ])

    # Simpan ke memory
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, as_attachment=True, download_name="data_santri.xlsx", mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# List Jenis Pembayaran
@app.route('/jenis_pembayaran')
def list_jenis_pembayaran():
    if 'username' not in session:
        return redirect(url_for('login'))

    conn = get_db_connection()
    jenis_list = conn.execute('SELECT * FROM jenis_pembayaran').fetchall()
    conn.close()
    return render_template('list_jenis_pembayaran.html', jenis_list=jenis_list)

# Tambah Jenis Pembayaran
@app.route('/tambah_jenis_pembayaran', methods=['GET', 'POST'])
def tambah_jenis_pembayaran():
    if 'username' not in session:
        return redirect(url_for('login'))

    if request.method == 'POST':
        nama_pembayaran = request.form['nama_pembayaran']
        nominal_default = request.form['nominal_default']

        conn = get_db_connection()
        conn.execute('INSERT INTO jenis_pembayaran (nama_pembayaran, nominal_default) VALUES (?, ?)',
                     (nama_pembayaran, nominal_default))
        conn.commit()
        conn.close()

        flash('Jenis pembayaran berhasil ditambahkan.', 'success')
        return redirect(url_for('list_jenis_pembayaran'))

    return render_template('tambah_jenis_pembayaran.html')

# Edit Jenis Pembayaran
@app.route('/edit_jenis_pembayaran/<int:id>', methods=['GET', 'POST'])
def edit_jenis_pembayaran(id):
    if 'username' not in session:
        return redirect(url_for('login'))

    conn = get_db_connection()
    jenis = conn.execute('SELECT * FROM jenis_pembayaran WHERE id_jenis = ?', (id,)).fetchone()

    if request.method == 'POST':
        nama_pembayaran = request.form['nama_pembayaran']
        nominal_default = request.form['nominal_default']

        conn.execute('UPDATE jenis_pembayaran SET nama_pembayaran = ?, nominal_default = ? WHERE id_jenis = ?',
                     (nama_pembayaran, nominal_default, id))
        conn.commit()
        conn.close()

        flash('Jenis pembayaran berhasil diperbarui.', 'success')
        return redirect(url_for('list_jenis_pembayaran'))

    conn.close()
    return render_template('edit_jenis_pembayaran.html', jenis=jenis)

# Hapus Jenis Pembayaran
@app.route('/hapus_jenis_pembayaran/<int:id>')
def hapus_jenis_pembayaran(id):
    if 'username' not in session:
        return redirect(url_for('login'))

    conn = get_db_connection()
    conn.execute('DELETE FROM jenis_pembayaran WHERE id_jenis = ?', (id,))
    conn.commit()
    conn.close()

    flash('Jenis pembayaran berhasil dihapus.', 'success')
    return redirect(url_for('list_jenis_pembayaran'))

# Tambahan route untuk Tambah Pembayaran di app.py

@app.route('/tambah_pembayaran', methods=['GET', 'POST'])
def tambah_pembayaran():
    if 'username' not in session:
        return redirect(url_for('login'))

    conn = get_db_connection()
    santri_list = conn.execute('SELECT * FROM santri').fetchall()
    jenis_list = conn.execute('SELECT * FROM jenis_pembayaran').fetchall()

    if request.method == 'POST':
        id_santri = request.form['id_santri']
        jenis_pembayaran = request.form['jenis_pembayaran']
        bulan = request.form['bulan']
        tahun = request.form['tahun']
        nominal_tagihan = request.form['nominal_tagihan']
        nominal_dibayar = request.form['nominal_dibayar']
        tanggal_bayar = request.form['tanggal_bayar']
        metode_bayar = request.form['metode_bayar']
        status_bayar = request.form['status_bayar']
        catatan = request.form['catatan']
        kasir = session['username']

        conn.execute('''
            INSERT INTO pembayaran (
                id_santri, jenis_pembayaran, bulan, tahun, nominal_tagihan,
                nominal_dibayar, tanggal_bayar, metode_bayar, status_bayar, kasir, catatan
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (id_santri, jenis_pembayaran, bulan, tahun, nominal_tagihan,
              nominal_dibayar, tanggal_bayar, metode_bayar, status_bayar, kasir, catatan))
        conn.commit()
        conn.close()

        flash('Pembayaran berhasil disimpan.', 'success')
        return redirect(url_for('dashboard'))

    conn.close()
    return render_template('tambah_pembayaran.html', santri_list=santri_list, jenis_list=jenis_list)

# Tambahan route untuk List Pembayaran di app.py

@app.route('/list_pembayaran', methods=['GET', 'POST'])
def list_pembayaran():
    if 'username' not in session:
        return redirect(url_for('login'))

    conn = get_db_connection()

    bulan = request.args.get('bulan')
    tahun = request.args.get('tahun')
    query = request.args.get('q')  # menangkap kata kunci pencarian

    sql = '''
        SELECT pembayaran.*, santri.nama_lengkap, santri.nis
        FROM pembayaran 
        JOIN santri ON pembayaran.id_santri = santri.id_santri
    '''
    params = []

    filters = []
    if bulan:
        filters.append("pembayaran.bulan = ?")
        params.append(bulan)
    if tahun:
        filters.append("pembayaran.tahun = ?")
        params.append(tahun)
    if query:
        filters.append("(santri.nama_lengkap LIKE ? OR santri.nis LIKE ?)")
        params.extend([f'%{query}%', f'%{query}%'])

    if filters:
        sql += " WHERE " + " AND ".join(filters)

    sql += " ORDER BY pembayaran.tanggal_bayar DESC"

    pembayaran_list = conn.execute(sql, params).fetchall()
    conn.close()

    return render_template(
        'list_pembayaran.html',
        pembayaran_list=pembayaran_list,
        bulan=bulan,
        tahun=tahun,
        query=query
    )

@app.route('/export_pembayaran')
def export_pembayaran():
    if 'username' not in session:
        return redirect(url_for('login'))

    conn = get_db_connection()
    pembayaran_list = conn.execute('''
        SELECT pembayaran.*, santri.nama_lengkap 
        FROM pembayaran 
        JOIN santri ON pembayaran.id_santri = santri.id_santri
        ORDER BY pembayaran.tanggal_bayar DESC
    ''').fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Data Pembayaran"

    # Header kolom
    headers = [
        'Nama Santri', 'Jenis Pembayaran', 'Bulan - Tahun', 'Nominal Tagihan', 'Nominal Dibayar',
        'Tanggal Bayar', 'Metode Bayar', 'Status Bayar', 'Kasir', 'Catatan'
    ]
    ws.append(headers)

    # Isi data
    for pembayaran in pembayaran_list:
        ws.append([
            pembayaran['nama_lengkap'],
            pembayaran['jenis_pembayaran'],
            f"{pembayaran['bulan']} {pembayaran['tahun']}",
            pembayaran['nominal_tagihan'],
            pembayaran['nominal_dibayar'],
            pembayaran['tanggal_bayar'],
            pembayaran['metode_bayar'],
            pembayaran['status_bayar'],
            pembayaran['kasir'],
            pembayaran['catatan']
        ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, as_attachment=True, download_name="data_pembayaran.xlsx", mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/detail_pembayaran/<int:id>')
def detail_pembayaran(id):
    if 'username' not in session:
        return redirect(url_for('login'))

    conn = get_db_connection()
    pembayaran = conn.execute('''
        SELECT pembayaran.*, santri.nama_lengkap 
        FROM pembayaran 
        JOIN santri ON pembayaran.id_santri = santri.id_santri
        WHERE pembayaran.id_pembayaran = ?
    ''', (id,)).fetchone()
    conn.close()

    if pembayaran is None:
        flash('Data pembayaran tidak ditemukan.', 'danger')
        return redirect(url_for('list_pembayaran'))

    return render_template('detail_pembayaran.html', pembayaran=pembayaran)

# Tambahan route untuk List Tunggakan di app.py

@app.route('/list_tunggakan', methods=['GET', 'POST'])
def list_tunggakan():
    if 'username' not in session:
        return redirect(url_for('login'))

    conn = get_db_connection()

    bulan = request.args.get('bulan')
    tahun = request.args.get('tahun')

    if bulan and tahun:
        tunggakan_list = conn.execute('''
            SELECT pembayaran.*, santri.nama_lengkap 
            FROM pembayaran 
            JOIN santri ON pembayaran.id_santri = santri.id_santri
            WHERE pembayaran.status_bayar = 'Belum Lunas' AND pembayaran.bulan = ? AND pembayaran.tahun = ?
            ORDER BY pembayaran.tanggal_bayar DESC
        ''', (bulan, tahun)).fetchall()
    else:
        tunggakan_list = conn.execute('''
            SELECT pembayaran.*, santri.nama_lengkap 
            FROM pembayaran 
            JOIN santri ON pembayaran.id_santri = santri.id_santri
            WHERE pembayaran.status_bayar = 'Belum Lunas'
            ORDER BY pembayaran.tanggal_bayar DESC
        ''').fetchall()

    conn.close()

    return render_template('list_tunggakan.html', tunggakan_list=tunggakan_list, bulan=bulan, tahun=tahun)

@app.route('/export_tunggakan')
def export_tunggakan():
    if 'username' not in session:
        return redirect(url_for('login'))

    conn = get_db_connection()
    tunggakan_list = conn.execute('''
        SELECT pembayaran.*, santri.nama_lengkap 
        FROM pembayaran 
        JOIN santri ON pembayaran.id_santri = santri.id_santri
        WHERE pembayaran.status_bayar = 'Belum Lunas'
        ORDER BY pembayaran.tanggal_bayar DESC
    ''').fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Data Tunggakan"

    # Header kolom
    headers = [
        'Nama Santri', 'Jenis Pembayaran', 'Bulan - Tahun', 'Nominal Tagihan', 'Nominal Dibayar', 'Kekurangan', 'Status Bayar'
    ]
    ws.append(headers)

    # Isi data
    for pembayaran in tunggakan_list:
        ws.append([
            pembayaran['nama_lengkap'],
            pembayaran['jenis_pembayaran'],
            f"{pembayaran['bulan']} {pembayaran['tahun']}",
            pembayaran['nominal_tagihan'],
            pembayaran['nominal_dibayar'],
            pembayaran['nominal_tagihan'] - pembayaran['nominal_dibayar'],
            pembayaran['status_bayar']
        ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, as_attachment=True, download_name="data_tunggakan.xlsx", mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# Tambahan route untuk Laporan Keuangan Per Bulan di app.py

@app.route('/laporan_bulanan', methods=['GET', 'POST'])
def laporan_bulanan():
    if 'username' not in session:
        return redirect(url_for('login'))

    bulan = request.args.get('bulan')
    tahun = request.args.get('tahun')

    laporan_list = []
    total_bulan_ini = 0

    if bulan and tahun:
        conn = get_db_connection()
        laporan_list = conn.execute('''
            SELECT pembayaran.*, santri.nama_lengkap 
            FROM pembayaran 
            JOIN santri ON pembayaran.id_santri = santri.id_santri
            WHERE pembayaran.bulan = ? AND pembayaran.tahun = ?
            ORDER BY pembayaran.tanggal_bayar ASC
        ''', (bulan, tahun)).fetchall()

        total_bulan_ini = conn.execute('''
            SELECT IFNULL(SUM(nominal_dibayar),0) FROM pembayaran
            WHERE bulan = ? AND tahun = ?
        ''', (bulan, tahun)).fetchone()[0]

        conn.close()

    return render_template('laporan_bulanan.html', laporan_list=laporan_list, bulan=bulan, tahun=tahun, total_bulan_ini=total_bulan_ini)

@app.route('/export_laporan_bulanan')
def export_laporan_bulanan():
    if 'username' not in session:
        return redirect(url_for('login'))

    bulan = request.args.get('bulan')
    tahun = request.args.get('tahun')

    if not bulan or not tahun:
        flash('Pilih Bulan dan Tahun terlebih dahulu untuk export.', 'warning')
        return redirect(url_for('laporan_bulanan'))

    conn = get_db_connection()
    laporan_list = conn.execute('''
        SELECT pembayaran.*, santri.nama_lengkap 
        FROM pembayaran 
        JOIN santri ON pembayaran.id_santri = santri.id_santri
        WHERE pembayaran.bulan = ? AND pembayaran.tahun = ?
        ORDER BY pembayaran.tanggal_bayar ASC
    ''', (bulan, tahun)).fetchall()

    total_bulan_ini = conn.execute('''
        SELECT IFNULL(SUM(nominal_dibayar),0) FROM pembayaran
        WHERE bulan = ? AND tahun = ?
    ''', (bulan, tahun)).fetchone()[0]
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan Bulanan"

    # Header kolom
    headers = ['Nama Santri', 'Jenis Pembayaran', 'Nominal Dibayar', 'Tanggal Bayar', 'Metode Pembayaran']
    ws.append(headers)

    # Isi data
    for pembayaran in laporan_list:
        ws.append([
            pembayaran['nama_lengkap'],
            pembayaran['jenis_pembayaran'],
            pembayaran['nominal_dibayar'],
            pembayaran['tanggal_bayar'],
            pembayaran['metode_bayar']
        ])

    # Tambahkan total di bawah
    ws.append([])
    ws.append(['', '', 'TOTAL', total_bulan_ini, ''])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"laporan_{bulan}_{tahun}.xlsx"

    return send_file(output, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# Tambahan route Rekap Keuangan Tahunan di app.py

@app.route('/rekap_tahunan', methods=['GET'])
def rekap_tahunan():
    if 'username' not in session:
        return redirect(url_for('login'))

    tahun = request.args.get('tahun')
    rekap_list = []
    total_pemasukan = 0
    total_tunggakan = 0
    bulan_labels = []
    total_values = []

    if tahun:
        conn = get_db_connection()
        rekap_list = conn.execute('''
            SELECT bulan,
                SUM(nominal_dibayar) AS total_pemasukan,
                SUM(nominal_tagihan - nominal_dibayar) AS total_tunggakan
            FROM pembayaran
            WHERE tahun = ?
            GROUP BY bulan
            ORDER BY CASE
                WHEN bulan = 'Januari' THEN 1
                WHEN bulan = 'Februari' THEN 2
                WHEN bulan = 'Maret' THEN 3
                WHEN bulan = 'April' THEN 4
                WHEN bulan = 'Mei' THEN 5
                WHEN bulan = 'Juni' THEN 6
                WHEN bulan = 'Juli' THEN 7
                WHEN bulan = 'Agustus' THEN 8
                WHEN bulan = 'September' THEN 9
                WHEN bulan = 'Oktober' THEN 10
                WHEN bulan = 'November' THEN 11
                WHEN bulan = 'Desember' THEN 12
            END
        ''', (tahun,)).fetchall()

        total_pemasukan = conn.execute('''
            SELECT IFNULL(SUM(nominal_dibayar),0)
            FROM pembayaran
            WHERE tahun = ?
        ''', (tahun,)).fetchone()[0]

        total_tunggakan = conn.execute('''
            SELECT IFNULL(SUM(nominal_tagihan - nominal_dibayar),0)
            FROM pembayaran
            WHERE tahun = ?
        ''', (tahun,)).fetchone()[0]

        for item in rekap_list:
            bulan_labels.append(item['bulan'])
            total_values.append(item['total_pemasukan'])

        conn.close()

    return render_template('rekap_tahunan.html', 
                           rekap_list=rekap_list, tahun=tahun, 
                           total_pemasukan=total_pemasukan, 
                           total_tunggakan=total_tunggakan,
                           bulan_labels=bulan_labels,
                           total_values=total_values)

# Tambahan route Export Rekap Tahunan di app.py

@app.route('/export_rekap_tahunan')
def export_rekap_tahunan():
    if 'username' not in session:
        return redirect(url_for('login'))

    tahun = request.args.get('tahun')

    if not tahun:
        flash('Pilih Tahun terlebih dahulu untuk export.', 'warning')
        return redirect(url_for('rekap_tahunan'))

    conn = get_db_connection()
    rekap_list = conn.execute('''
        SELECT bulan,
            SUM(nominal_dibayar) AS total_pemasukan,
            SUM(nominal_tagihan - nominal_dibayar) AS total_tunggakan
        FROM pembayaran
        WHERE tahun = ?
        GROUP BY bulan
        ORDER BY CASE
            WHEN bulan = 'Januari' THEN 1
            WHEN bulan = 'Februari' THEN 2
            WHEN bulan = 'Maret' THEN 3
            WHEN bulan = 'April' THEN 4
            WHEN bulan = 'Mei' THEN 5
            WHEN bulan = 'Juni' THEN 6
            WHEN bulan = 'Juli' THEN 7
            WHEN bulan = 'Agustus' THEN 8
            WHEN bulan = 'September' THEN 9
            WHEN bulan = 'Oktober' THEN 10
            WHEN bulan = 'November' THEN 11
            WHEN bulan = 'Desember' THEN 12
        END
    ''', (tahun,)).fetchall()

    total_pemasukan = conn.execute('''
        SELECT IFNULL(SUM(nominal_dibayar),0)
        FROM pembayaran
        WHERE tahun = ?
    ''', (tahun,)).fetchone()[0]

    total_tunggakan = conn.execute('''
        SELECT IFNULL(SUM(nominal_tagihan - nominal_dibayar),0)
        FROM pembayaran
        WHERE tahun = ?
    ''', (tahun,)).fetchone()[0]
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Rekap Tahunan"

    # Header kolom
    headers = ['Bulan', 'Total Pemasukan', 'Total Tunggakan']
    ws.append(headers)

    # Isi data
    for item in rekap_list:
        ws.append([
            item['bulan'],
            item['total_pemasukan'],
            item['total_tunggakan']
        ])

    # Tambahkan total di bawah
    ws.append([])
    ws.append(['TOTAL', total_pemasukan, total_tunggakan])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"rekap_{tahun}.xlsx"

    return send_file(output, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/backup_database')
def backup_database():
    try:
        # Pastikan folder backup ada
        backup_folder = 'backup'
        if not os.path.exists(backup_folder):
            os.makedirs(backup_folder)

        # Tentukan nama file backup
        database_file = 'erp_pondok.db'
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        backup_filename = f'backup_{timestamp}.db'
        backup_path = os.path.join(backup_folder, backup_filename)

        # Salin database ke backup
        shutil.copy2(database_file, backup_path)

        # Siapkan file untuk langsung di-download
        return send_file(
            backup_path,
            as_attachment=True,
            download_name=backup_filename,
            mimetype='application/octet-stream'
        )
    except Exception as e:
        flash(f'Gagal backup database: {str(e)}', 'danger')
        return redirect(url_for('dashboard'))

@app.route('/list_backup')
def list_backup():
    if 'username' not in session:
        return redirect(url_for('login'))

    backup_folder = 'backup'
    backups = []

    if os.path.exists(backup_folder):
        for filename in os.listdir(backup_folder):
            if filename.endswith('.db'):
                full_path = os.path.join(backup_folder, filename)
                backups.append({
                    'filename': filename,
                    'size_kb': round(os.path.getsize(full_path) / 1024, 2),
                    'modified': datetime.fromtimestamp(os.path.getmtime(full_path)).strftime('%Y-%m-%d %H:%M:%S')
                })

    backups = sorted(backups, key=lambda x: x['modified'], reverse=True)
    return render_template('list_backup.html', backups=backups)

@app.route('/download_backup/<filename>')
def download_backup(filename):
    backup_path = os.path.join('backup', filename)
    if os.path.exists(backup_path):
        return send_file(backup_path, as_attachment=True)
    else:
        flash('File backup tidak ditemukan.', 'danger')
        return redirect(url_for('list_backup'))

@app.route('/hapus_backup/<filename>')
def hapus_backup(filename):
    backup_path = os.path.join('backup', filename)
    if os.path.exists(backup_path):
        os.remove(backup_path)
        flash('File backup berhasil dihapus.', 'success')
    else:
        flash('File backup tidak ditemukan.', 'danger')
    return redirect(url_for('list_backup'))

@app.route('/restore_backup/<filename>')
def restore_backup(filename):
    try:
        backup_path = os.path.join('backup', filename)
        database_path = 'erp_pondok.db'

        if os.path.exists(backup_path):
            shutil.copy2(backup_path, database_path)
            flash(f'Database berhasil direstore dari {filename}', 'success')
        else:
            flash('File backup tidak ditemukan.', 'danger')
    except Exception as e:
        flash(f'Gagal restore database: {str(e)}', 'danger')

    return redirect(url_for('list_backup'))

@app.route('/settings', methods=['GET', 'POST'])
def settings_aplikasi():
    if 'username' not in session or session.get('role') != 'admin':
        return redirect(url_for('login'))

    if request.method == 'POST':
        data = request.form.to_dict()
        if 'logo' in request.files:
            logo_file = request.files['logo']
            if logo_file and logo_file.filename:
                logo_path = os.path.join('static/uploads', logo_file.filename)
                logo_file.save(logo_path)
                data['logo'] = logo_file.filename
        else:
            data['logo'] = load_settings()['logo'] if load_settings() else ''

        save_settings(data)
        flash('Pengaturan berhasil disimpan.', 'success')
        return redirect(url_for('settings'))

    settings = load_settings()
    return render_template('settings.html', settings=settings)

@app.route('/cari_santri')
def cari_santri():
    keyword = request.args.get('q', '')
    conn = get_db_connection()
    results = conn.execute('''
        SELECT id_santri, nama_lengkap, nis 
        FROM santri
        WHERE nama_lengkap LIKE ? OR nis LIKE ?
        LIMIT 10
    ''', (f'%{keyword}%', f'%{keyword}%')).fetchall()
    conn.close()
    return jsonify([{'id': r['id_santri'], 'label': f"{r['nis']} - {r['nama_lengkap']}"} for r in results])

@app.route('/import_santri', methods=['GET', 'POST'])
def import_santri():
    if request.method == 'POST':
        file = request.files['file']
        if file:
            filename = secure_filename(file.filename)
            filepath = os.path.join('uploads', filename)
            os.makedirs('uploads', exist_ok=True)  # buat folder kalau belum ada
            file.save(filepath)

            # Mengecek apakah file yang di-upload adalah CSV atau Excel
            if filename.endswith('.csv'):
                df = pd.read_csv(filepath)
            elif filename.endswith('.xls') or filename.endswith('.xlsx'):
                df = pd.read_excel(filepath)
            else:
                flash('File yang di-upload bukan file CSV atau Excel.')
                return redirect(url_for('import_santri'))

            # Menambahkan data ke dalam database
            for index, row in df.iterrows():
                # Cek apakah 'nis' sudah ada di database
                existing_santri = Santri.query.filter_by(nis=row['nis']).first()
                if existing_santri:
                    continue  # Skip jika sudah ada

                # Mengonversi tanggal menjadi format yang benar
                try:
                    tanggal_lahir = datetime.strptime(row['tanggal_lahir'], '%Y-%m-%d').date()
                    tanggal_masuk = datetime.strptime(row['tanggal_masuk'], '%Y-%m-%d').date()
                except ValueError:
                    tanggal_lahir = None
                    tanggal_masuk = None

                # Membuat objek Santri dan menambahkannya ke session
                santri = Santri(
                    nis=row['nis'],
                    nama_lengkap=row['nama_lengkap'],
                    tempat_lahir=row['tempat_lahir'],
                    tanggal_lahir=tanggal_lahir,
                    jenis_kelamin=row['jenis_kelamin'],
                    alamat=row['alamat'],
                    no_hp=row['no_hp'],
                    wali_santri=row['wali_santri'],
                    kelas=row['kelas'],
                    kamar=row['kamar'],
                    status=row['status'],
                    tanggal_masuk=tanggal_masuk
                )
                db.session.add(santri)

            db.session.commit()

            return 'Data berhasil diimpor'
    return render_template('import_santri.html')
@app.route('/download_template_santri')
def download_template_santri():
    return send_file('template_santri.xlsx', as_attachment=True)

if __name__ == '__main__':
    app.run(debug=True)
